"""Tests for building multi-worksheet audit workbook packages."""

from pathlib import Path

from openpyxl import Workbook

from auditor_support_tool.core.workbook_package import (
    DatasetType,
    MappingConfidence,
)
from auditor_support_tool.core.workbook_package_service import (
    WorkbookPackageService,
)


def create_workbook(tmp_path: Path) -> Path:
    """Create a workbook containing several audit datasets."""

    workbook = Workbook()

    general_ledger = workbook.active
    general_ledger.title = "General_Ledger"
    general_ledger.append(
        [
            "Transaction Date",
            "Account Code",
            "Invoice Number",
            "Amount",
        ]
    )
    general_ledger.append(
        [
            "2026-01-01",
            "1000",
            "INV-001",
            100.00,
        ]
    )

    chart = workbook.create_sheet("Chart_of_Accounts")
    chart.append(
        [
            "Account Code",
            "Account Name",
            "Account Type",
        ]
    )
    chart.append(
        [
            "1000",
            "Cash",
            "Asset",
        ]
    )

    instructions = workbook.create_sheet("Instructions")
    instructions.append(["Instruction"])
    instructions.append(["Do not import this sheet automatically."])

    workbook.create_sheet("Empty")

    path = tmp_path / "audit_data.xlsx"
    workbook.save(path)

    return path


def test_package_loads_all_non_empty_worksheets(
    tmp_path: Path,
) -> None:
    """Each non-empty worksheet should become a dataset."""

    path = create_workbook(tmp_path)

    package = WorkbookPackageService().build_package(path)

    assert package.source_file_name == "audit_data.xlsx"

    assert tuple(dataset.original_worksheet_name for dataset in package.datasets) == (
        "General_Ledger",
        "Chart_of_Accounts",
        "Instructions",
    )


def test_empty_worksheet_is_excluded_by_default(
    tmp_path: Path,
) -> None:
    """Completely empty worksheets should not be loaded by default."""

    path = create_workbook(tmp_path)

    package = WorkbookPackageService().build_package(path)

    assert package.get_dataset_by_worksheet("Empty") is None


def test_general_ledger_is_suggested(
    tmp_path: Path,
) -> None:
    """The General Ledger sheet should receive an editable suggestion."""

    path = create_workbook(tmp_path)

    package = WorkbookPackageService().build_package(path)
    dataset = package.get_dataset_by_worksheet("General_Ledger")

    assert dataset is not None
    assert dataset.suggested_dataset_type == DatasetType.GENERAL_LEDGER
    assert dataset.confirmed_dataset_type == DatasetType.GENERAL_LEDGER
    assert dataset.suggested_display_name == "General Ledger"
    assert dataset.suggestion_confidence in {
        MappingConfidence.MEDIUM,
        MappingConfidence.HIGH,
    }


def test_chart_of_accounts_is_suggested(
    tmp_path: Path,
) -> None:
    """The Chart of Accounts sheet should be recognised."""

    path = create_workbook(tmp_path)

    package = WorkbookPackageService().build_package(path)
    dataset = package.get_dataset_by_worksheet("Chart_of_Accounts")

    assert dataset is not None
    assert dataset.confirmed_dataset_type == DatasetType.CHART_OF_ACCOUNTS
    assert dataset.confirmed_display_name == "Chart of Accounts"


def test_columns_retain_detected_and_confirmed_types(
    tmp_path: Path,
) -> None:
    """Column suggestions should remain editable and traceable."""

    path = create_workbook(tmp_path)

    package = WorkbookPackageService().build_package(path)
    dataset = package.get_dataset_by_worksheet("General_Ledger")

    assert dataset is not None
    assert len(dataset.columns) == 4

    amount_column = next(column for column in dataset.columns if column.source_column == "Amount")

    assert amount_column.confirmed_type == amount_column.detected_type
    assert amount_column.confirmed_name == "Amount"


def test_original_worksheet_name_is_preserved(
    tmp_path: Path,
) -> None:
    """Changing suggestions must not affect source traceability."""

    path = create_workbook(tmp_path)

    package = WorkbookPackageService().build_package(path)
    dataset = package.get_dataset_by_worksheet("General_Ledger")

    assert dataset is not None

    dataset.confirmed_display_name = "2026 General Ledger"
    dataset.confirmed_dataset_type = DatasetType.OTHER

    assert dataset.original_worksheet_name == "General_Ledger"
    assert dataset.confirmed_display_name == "2026 General Ledger"
    assert dataset.confirmed_dataset_type == DatasetType.OTHER
