"""Tests for the General Ledger data-import foundation."""

from pathlib import Path

import pytest
from openpyxl import Workbook

from auditor_support_tool.core.data_import_service import (
    CSV_WORKSHEET_NAME,
    DataImportError,
    DataImportService,
)
from auditor_support_tool.core.data_models import (
    SOURCE_ROW_FIELD,
)


@pytest.fixture
def service() -> DataImportService:
    """Return a data-import service for each test."""

    return DataImportService()


@pytest.fixture
def sample_excel_file(tmp_path: Path) -> Path:
    """Create a small workbook containing representative audit data."""

    workbook = Workbook()

    general_ledger = workbook.active
    general_ledger.title = "General_Ledger"

    general_ledger.append(
        [
            "Transaction Date",
            "Invoice Number",
            "Vendor Number",
            "Amount",
        ]
    )
    general_ledger.append(
        [
            "2026-01-05",
            "INV-001",
            "V001",
            100.00,
        ]
    )
    general_ledger.append(
        [
            "2026-01-06",
            "INV-002",
            "V002",
            250.00,
        ]
    )
    general_ledger.append([None, None, None, None])
    general_ledger.append(
        [
            "2026-01-07",
            "INV-003",
            "V003",
            300.00,
        ]
    )

    supporting_sheet = workbook.create_sheet("Supporting_Data")
    supporting_sheet.append(["Reference", "Description"])
    supporting_sheet.append(["A1", "Supporting record"])

    path = tmp_path / "sample_general_ledger.xlsx"
    workbook.save(path)

    return path


@pytest.fixture
def unusual_headers_file(tmp_path: Path) -> Path:
    """Create a workbook with blank, duplicate and reserved headers."""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "General_Ledger"

    worksheet.append(
        [
            "Invoice Number",
            "",
            "Invoice Number",
            SOURCE_ROW_FIELD,
        ]
    )
    worksheet.append(
        [
            "INV-001",
            "Supporting value",
            "INV-001",
            "Original source value",
        ]
    )

    path = tmp_path / "unusual_headers.xlsx"
    workbook.save(path)

    return path


@pytest.fixture
def sample_csv_file(tmp_path: Path) -> Path:
    """Create a small CSV population."""

    path = tmp_path / "sample_general_ledger.csv"
    path.write_text(
        "Transaction Date,Invoice Number,Amount\n"
        "2026-01-05,INV-001,100.00\n"
        "2026-01-06,INV-002,250.00\n",
        encoding="utf-8",
    )

    return path


def test_inspect_excel_source_lists_worksheets(
    service: DataImportService,
    sample_excel_file: Path,
) -> None:
    """Excel inspection should return all workbook worksheets."""

    source = service.inspect_source(sample_excel_file)

    assert source.file_type == "xlsx"
    assert source.path == sample_excel_file.resolve()
    assert source.file_size_bytes > 0

    assert len(source.worksheets) == 2
    assert source.worksheets[0].name == "General_Ledger"
    assert source.worksheets[0].position == 1
    assert source.worksheets[1].name == "Supporting_Data"
    assert source.worksheets[1].position == 2


def test_load_excel_population_preserves_source_rows(
    service: DataImportService,
    sample_excel_file: Path,
) -> None:
    """Loaded records should retain their original worksheet row numbers."""

    table = service.load_table(
        sample_excel_file,
        worksheet_name="General_Ledger",
    )

    assert table.worksheet_name == "General_Ledger"
    assert table.record_count == 3
    assert table.column_count == 4

    assert table.summary.source_records_read == 4
    assert table.summary.records_loaded == 3
    assert table.summary.blank_rows_skipped == 1

    assert table.rows[0][SOURCE_ROW_FIELD] == 2
    assert table.rows[1][SOURCE_ROW_FIELD] == 3
    assert table.rows[2][SOURCE_ROW_FIELD] == 5


def test_load_excel_population_preserves_headers_and_values(
    service: DataImportService,
    sample_excel_file: Path,
) -> None:
    """Headers and source values should remain available to later tests."""

    table = service.load_table(
        sample_excel_file,
        worksheet_name="General_Ledger",
    )

    assert table.headers == (
        "Transaction Date",
        "Invoice Number",
        "Vendor Number",
        "Amount",
    )

    assert table.rows[0]["Invoice Number"] == "INV-001"
    assert table.rows[0]["Vendor Number"] == "V001"
    assert table.rows[0]["Amount"] == 100.00


def test_blank_duplicate_and_reserved_headers_are_resolved(
    service: DataImportService,
    unusual_headers_file: Path,
) -> None:
    """Unsafe source headers should be resolved without losing data."""

    table = service.load_table(
        unusual_headers_file,
        worksheet_name="General_Ledger",
    )

    assert table.headers == (
        "Invoice Number",
        "Unnamed Column 2",
        "Invoice Number [2]",
        f"{SOURCE_ROW_FIELD} [source]",
    )

    assert len(table.summary.header_changes) == 3

    reasons = {change.reason for change in table.summary.header_changes}

    assert reasons == {
        "Blank header",
        "Duplicate header",
        "Reserved header",
    }

    assert table.rows[0]["Invoice Number"] == "INV-001"
    assert table.rows[0]["Unnamed Column 2"] == "Supporting value"
    assert table.rows[0]["Invoice Number [2]"] == "INV-001"
    assert table.rows[0][f"{SOURCE_ROW_FIELD} [source]"] == "Original source value"
    assert table.rows[0][SOURCE_ROW_FIELD] == 2


def test_inspect_csv_source_returns_single_logical_worksheet(
    service: DataImportService,
    sample_csv_file: Path,
) -> None:
    """CSV inspection should expose one logical data source."""

    source = service.inspect_source(sample_csv_file)

    assert source.file_type == "csv"
    assert len(source.worksheets) == 1
    assert source.worksheets[0].name == CSV_WORKSHEET_NAME
    assert source.worksheets[0].position == 1
    assert source.worksheets[0].estimated_data_rows == 2


def test_load_csv_population(
    service: DataImportService,
    sample_csv_file: Path,
) -> None:
    """CSV data should use the same loaded-table structure as Excel."""

    table = service.load_table(sample_csv_file)

    assert table.file_type == "csv"
    assert table.worksheet_name == CSV_WORKSHEET_NAME
    assert table.record_count == 2
    assert table.column_count == 3

    assert table.rows[0]["Invoice Number"] == "INV-001"
    assert table.rows[0][SOURCE_ROW_FIELD] == 2
    assert table.rows[1][SOURCE_ROW_FIELD] == 3


def test_missing_worksheet_raises_clear_error(
    service: DataImportService,
    sample_excel_file: Path,
) -> None:
    """A missing worksheet should not silently fall back to another sheet."""

    with pytest.raises(
        DataImportError,
        match="Worksheet not found",
    ):
        service.load_table(
            sample_excel_file,
            worksheet_name="Missing_Sheet",
        )


def test_unsupported_file_type_is_rejected(
    service: DataImportService,
    tmp_path: Path,
) -> None:
    """Unsupported source types should be rejected before processing."""

    unsupported_file = tmp_path / "general_ledger.txt"
    unsupported_file.write_text(
        "Not a supported audit data source.",
        encoding="utf-8",
    )

    with pytest.raises(
        DataImportError,
        match="Unsupported file type",
    ):
        service.inspect_source(unsupported_file)


def test_missing_source_file_raises_clear_error(
    service: DataImportService,
    tmp_path: Path,
) -> None:
    """A missing source should produce a specific import error."""

    missing_file = tmp_path / "missing.xlsx"

    with pytest.raises(
        DataImportError,
        match="Source file not found",
    ):
        service.inspect_source(missing_file)


def test_invalid_header_row_is_rejected(
    service: DataImportService,
    sample_excel_file: Path,
) -> None:
    """Header rows must use one-based worksheet numbering."""

    with pytest.raises(
        ValueError,
        match="header row number must be at least 1",
    ):
        service.load_table(
            sample_excel_file,
            worksheet_name="General_Ledger",
            header_row=0,
        )
