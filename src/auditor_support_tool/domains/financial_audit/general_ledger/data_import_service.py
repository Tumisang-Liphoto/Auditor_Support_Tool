"""Excel and CSV import services for General Ledger audit analytics."""

import csv
from collections.abc import Iterable, Sequence
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from auditor_support_tool.domains.financial_audit.general_ledger.models import (
    SOURCE_ROW_FIELD,
    HeaderChange,
    LoadedTable,
    PopulationSummary,
    SourceFileInfo,
    WorksheetInfo,
)

SUPPORTED_EXTENSIONS = {
    ".xlsx",
    ".xlsm",
    ".csv",
}

CSV_WORKSHEET_NAME = "CSV Data"


class DataImportError(RuntimeError):
    """Raised when an audit data source cannot be inspected or loaded."""


class DataImportService:
    """Discover and load supported Excel and CSV data sources."""

    def inspect_source(
        self,
        source_path: str | Path,
    ) -> SourceFileInfo:
        """Return file and worksheet metadata without loading test data."""

        path = self._validate_source_path(source_path)
        extension = path.suffix.lower()

        if extension == ".csv":
            worksheet = self._inspect_csv(path)

            return SourceFileInfo(
                path=path,
                file_type="csv",
                file_size_bytes=path.stat().st_size,
                worksheets=(worksheet,),
            )

        worksheets = self._inspect_excel(path)

        return SourceFileInfo(
            path=path,
            file_type=extension.removeprefix("."),
            file_size_bytes=path.stat().st_size,
            worksheets=worksheets,
        )

    def load_table(
        self,
        source_path: str | Path,
        *,
        worksheet_name: str | None = None,
        header_row: int = 1,
    ) -> LoadedTable:
        """Load a selected worksheet or CSV population.

        Completely blank data rows are skipped. Every retained record includes
        the original Excel or CSV row number in ``_source_row_number``.
        """

        path = self._validate_source_path(source_path)

        if header_row < 1:
            raise ValueError("The header row number must be at least 1.")

        extension = path.suffix.lower()

        if extension == ".csv":
            return self._load_csv(
                path,
                header_row=header_row,
            )

        return self._load_excel(
            path,
            worksheet_name=worksheet_name,
            header_row=header_row,
        )

    def _inspect_excel(
        self,
        path: Path,
    ) -> tuple[WorksheetInfo, ...]:
        try:
            workbook = load_workbook(
                filename=path,
                read_only=True,
                data_only=True,
            )
        except Exception as error:
            raise DataImportError(f"Unable to inspect Excel workbook: {error}") from error

        try:
            worksheets = tuple(
                WorksheetInfo(
                    name=worksheet.title,
                    position=index,
                    maximum_row=worksheet.max_row,
                    maximum_column=worksheet.max_column,
                    estimated_data_rows=max(
                        worksheet.max_row - 1,
                        0,
                    ),
                )
                for index, worksheet in enumerate(
                    workbook.worksheets,
                    start=1,
                )
            )
        finally:
            workbook.close()

        if not worksheets:
            raise DataImportError("The Excel workbook does not contain any worksheets.")

        return worksheets

    def _inspect_csv(
        self,
        path: Path,
    ) -> WorksheetInfo:
        rows = self._read_csv_rows(path)

        maximum_row = len(rows)
        maximum_column = max(
            (len(row) for row in rows),
            default=0,
        )

        return WorksheetInfo(
            name=CSV_WORKSHEET_NAME,
            position=1,
            maximum_row=maximum_row,
            maximum_column=maximum_column,
            estimated_data_rows=max(maximum_row - 1, 0),
        )

    def _load_excel(
        self,
        path: Path,
        *,
        worksheet_name: str | None,
        header_row: int,
    ) -> LoadedTable:
        try:
            workbook = load_workbook(
                filename=path,
                read_only=True,
                data_only=True,
            )
        except Exception as error:
            raise DataImportError(f"Unable to open Excel workbook: {error}") from error

        try:
            if worksheet_name:
                if worksheet_name not in workbook.sheetnames:
                    raise DataImportError(f"Worksheet not found: {worksheet_name}")

                worksheet = workbook[worksheet_name]
            else:
                worksheet = workbook.worksheets[0]

            if header_row > worksheet.max_row:
                raise DataImportError("The selected header row is beyond the worksheet data.")

            row_iterator = worksheet.iter_rows(
                min_row=header_row,
                values_only=True,
            )

            try:
                raw_headers = next(row_iterator)
            except StopIteration as error:
                raise DataImportError("The selected worksheet is empty.") from error

            headers, original_headers, header_changes = self._resolve_headers(raw_headers)

            rows, summary_values = self._build_records(
                row_iterator,
                headers=headers,
                first_source_row=header_row + 1,
            )

            summary = PopulationSummary(
                source_records_read=summary_values["source_records_read"],
                records_loaded=len(rows),
                blank_rows_skipped=summary_values["blank_rows_skipped"],
                column_count=len(headers),
                blank_cell_count=summary_values["blank_cell_count"],
                header_changes=header_changes,
            )

            return LoadedTable(
                source_path=path,
                file_type=path.suffix.lower().removeprefix("."),
                worksheet_name=worksheet.title,
                headers=headers,
                original_headers=original_headers,
                rows=rows,
                summary=summary,
            )
        finally:
            workbook.close()

    def _load_csv(
        self,
        path: Path,
        *,
        header_row: int,
    ) -> LoadedTable:
        raw_rows = self._read_csv_rows(path)

        if not raw_rows:
            raise DataImportError("The selected CSV file is empty.")

        if header_row > len(raw_rows):
            raise DataImportError("The selected header row is beyond the CSV data.")

        raw_headers = raw_rows[header_row - 1]

        headers, original_headers, header_changes = self._resolve_headers(raw_headers)

        rows, summary_values = self._build_records(
            raw_rows[header_row:],
            headers=headers,
            first_source_row=header_row + 1,
        )

        summary = PopulationSummary(
            source_records_read=summary_values["source_records_read"],
            records_loaded=len(rows),
            blank_rows_skipped=summary_values["blank_rows_skipped"],
            column_count=len(headers),
            blank_cell_count=summary_values["blank_cell_count"],
            header_changes=header_changes,
        )

        return LoadedTable(
            source_path=path,
            file_type="csv",
            worksheet_name=CSV_WORKSHEET_NAME,
            headers=headers,
            original_headers=original_headers,
            rows=rows,
            summary=summary,
        )

    def _resolve_headers(
        self,
        raw_headers: Sequence[Any],
    ) -> tuple[
        tuple[str, ...],
        tuple[str, ...],
        tuple[HeaderChange, ...],
    ]:
        resolved_headers: list[str] = []
        original_headers: list[str] = []
        header_changes: list[HeaderChange] = []

        used_headers = {
            SOURCE_ROW_FIELD.casefold(),
        }

        for column_number, raw_header in enumerate(
            raw_headers,
            start=1,
        ):
            original = "" if raw_header is None else str(raw_header).strip()
            original_headers.append(original)

            if not original:
                base_header = f"Unnamed Column {column_number}"
                reason = "Blank header"
            elif original.casefold() == SOURCE_ROW_FIELD.casefold():
                base_header = f"{original} [source]"
                reason = "Reserved header"
            else:
                base_header = original
                reason = ""

            resolved = base_header
            suffix = 2

            while resolved.casefold() in used_headers:
                resolved = f"{base_header} [{suffix}]"
                suffix += 1

            if resolved != original:
                if not reason:
                    reason = "Duplicate header"

                header_changes.append(
                    HeaderChange(
                        column_number=column_number,
                        original_header=original,
                        resolved_header=resolved,
                        reason=reason,
                    )
                )

            used_headers.add(resolved.casefold())
            resolved_headers.append(resolved)

        return (
            tuple(resolved_headers),
            tuple(original_headers),
            tuple(header_changes),
        )

    def _build_records(
        self,
        raw_rows: Iterable[Sequence[Any]],
        *,
        headers: tuple[str, ...],
        first_source_row: int,
    ) -> tuple[
        tuple[dict[str, Any], ...],
        dict[str, int],
    ]:
        records: list[dict[str, Any]] = []

        source_records_read = 0
        blank_rows_skipped = 0
        blank_cell_count = 0

        for row_offset, raw_row in enumerate(raw_rows):
            source_records_read += 1
            source_row_number = first_source_row + row_offset

            values = list(raw_row[: len(headers)])

            if len(values) < len(headers):
                values.extend([None] * (len(headers) - len(values)))

            if self._is_blank_row(values):
                blank_rows_skipped += 1
                continue

            blank_cell_count += sum(1 for value in values if self._is_blank_value(value))

            record = {
                header: value
                for header, value in zip(
                    headers,
                    values,
                    strict=True,
                )
            }

            record[SOURCE_ROW_FIELD] = source_row_number
            records.append(record)

        return (
            tuple(records),
            {
                "source_records_read": source_records_read,
                "blank_rows_skipped": blank_rows_skipped,
                "blank_cell_count": blank_cell_count,
            },
        )

    def _read_csv_rows(
        self,
        path: Path,
    ) -> list[list[str]]:
        last_error: Exception | None = None

        for encoding in (
            "utf-8-sig",
            "utf-8",
            "cp1252",
        ):
            try:
                with path.open(
                    "r",
                    encoding=encoding,
                    newline="",
                ) as csv_file:
                    sample = csv_file.read(8192)
                    csv_file.seek(0)

                    try:
                        dialect = csv.Sniffer().sniff(
                            sample,
                            delimiters=",;\t|",
                        )
                    except csv.Error:
                        dialect = csv.excel

                    return [
                        list(row)
                        for row in csv.reader(
                            csv_file,
                            dialect,
                        )
                    ]
            except UnicodeDecodeError as error:
                last_error = error

        raise DataImportError(f"Unable to decode CSV file: {last_error}")

    @staticmethod
    def _validate_source_path(
        source_path: str | Path,
    ) -> Path:
        path = Path(source_path).expanduser().resolve()

        if not path.is_file():
            raise DataImportError(f"Source file not found: {path}")

        extension = path.suffix.lower()

        if extension not in SUPPORTED_EXTENSIONS:
            supported = ", ".join(sorted(SUPPORTED_EXTENSIONS))
            raise DataImportError(
                f"Unsupported file type '{extension}'. Supported types: {supported}"
            )

        return path

    @staticmethod
    def _is_blank_row(
        values: Sequence[Any],
    ) -> bool:
        return all(DataImportService._is_blank_value(value) for value in values)

    @staticmethod
    def _is_blank_value(
        value: Any,
    ) -> bool:
        if value is None:
            return True

        return isinstance(value, str) and not value.strip()
